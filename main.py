from openpyxl import load_workbook


EXCEL_FILE = 'resources/kala.xlsx'


def get_book(path):
    return load_workbook(path)


def get_sheet(excel_book):
    return excel_book.active


if __name__ == '__main__':
    book = get_book(EXCEL_FILE)

    sheet = get_sheet(book)

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        try:
            columns = [row[2], row[5], row[6]]
            cleaned_columns = []
            for column in columns:
                column_string = str(column)
                if column_string not in ['None', ' ', '']:
                    column_string_stripped = column_string.strip()
                    cleaned_columns.append(column_string_stripped)

            joined_columns = ' '.join(cleaned_columns)

        except Exception as e:
            print(e)
        else:

            sheet.cell(row=i, column=9).value = joined_columns

    book.save('./joined_kala.xlsx')
